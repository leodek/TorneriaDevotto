import tkinter as tk
from tkinter import ttk, messagebox
from tkcalendar import DateEntry
import sqlite3
from datetime import datetime
from ordenes_trabajo.ot_form_logic import abrir_formulario_modificar_ot
from ordenes_trabajo.agregar_orden_diaria import agregar_orden_trabajo
import os
import sys
import csv
from tkinter import filedialog, messagebox
import sqlite3
from db_init import get_db_path
from db_init import conectar



# Definición de campos y columnas al inicio del archivo
CAMPOS = [
    ("nro_ot", "Nº OT"),
    ("cliente", "Cliente"),
    ("trabajo", "Trabajo"),
    ("estado", "Estado"),
    ("prioridad", "Prioridad"),
    ("responsable", "Responsable"),
    ("observaciones", "Observaciones"),
    ("fecha_ingreso", "Fecha Ingreso"),
    ("fecha_estimada", "Fecha Estimada")
]

COLUMNAS = [campo[0] for campo in CAMPOS]

def mostrar_ordenes_trabajo(frame):
    """Muestra la interfaz de gestión de trabajos"""
    
    # Limpieza inicial del frame
    for widget in frame.winfo_children():
        widget.destroy()
    
    frame.configure(bg="white")
    frame.columnconfigure(0, weight=1)

    # Configuración del título
    titulo = tk.Label(
        frame,
        text="LISTA DE TRABAJO",
        font=("Arial", 20, "bold"),
        bg="white",
        fg="#3c3e50"
    )
    titulo.grid(row=0, column=0, pady=(20, 10))

    # Frame para botones superiores
    frame_botones = tk.Frame(frame, bg="white")
    frame_botones.grid(row=1, column=0, pady=(0, 10))

    # Función para cargar datos (definida antes de usarse)
    def cargar_datos():
        tree.delete(*tree.get_children())
        sql = f"SELECT {', '.join(COLUMNAS)} FROM ordenes_trabajo WHERE 1=1"
        params = []
        
        # Aplicar filtros
        for key, var in vars_filtro.items():
            if var.get():
                sql += f" AND {key} LIKE ?"
                params.append(f"%{var.get()}%")
        
        # Filtros de fecha
        if date_desde.get():
            try:
                fecha_desde = datetime.strptime(date_desde.get(), "%d/%m/%Y").strftime("%Y-%m-%d")
                sql += " AND date(fecha_ingreso) >= ?"
                params.append(fecha_desde)
            except ValueError:
                pass
                
        if date_hasta.get():
            try:
                fecha_hasta = datetime.strptime(date_hasta.get(), "%d/%m/%Y").strftime("%Y-%m-%d")
                sql += " AND date(fecha_ingreso) <= ?"
                params.append(fecha_hasta)
            except ValueError:
                pass

        # Ejecutar consulta
        try:
            with conectar() as conn:
                conn.row_factory = sqlite3.Row
                cur = conn.cursor()
                cur.execute(sql, params)
                for row in cur.fetchall():
                    tree.insert("", tk.END, values=[row[col] for col in COLUMNAS])
        except Exception as e:
            messagebox.showerror(
                "Error",
                f"No se pudo cargar los datos: {str(e)}",
                parent=frame_principal
            )

    def exportar_a_excel():
        # 1) Pedir ruta de guardado
        ruta = filedialog.asksaveasfilename(
            title="Guardar como...",
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("Excel 97-2003", "*.xls")],
            parent=frame
        )
        if not ruta:
            return

        # 2) Encabezados
        encabezados = [label for _, label in CAMPOS]

        try:
            # 3) Crear libro y estilos
            from openpyxl import Workbook
            from openpyxl.styles import Font, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            wb = Workbook()
            ws = wb.active

            # Definir estilos
            thin = Side(border_style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)
            center = Alignment(horizontal="center", vertical="center")

            # 4) Escribir encabezados (mayúscula, negrita, centrado, borde)
            for col_num, encabezado in enumerate(encabezados, start=1):
                cell = ws.cell(row=1, column=col_num, value=encabezado.upper())
                cell.font = Font(bold=True)
                cell.alignment = center
                cell.border = border

            # 5) Escribir datos (centrado y borde)
            for row_num, item_id in enumerate(tree.get_children(), start=2):
                valores = tree.item(item_id)["values"]
                for col_num, valor in enumerate(valores, start=1):
                    cell = ws.cell(row=row_num, column=col_num, value=valor)
                    cell.alignment = center
                    cell.border = border

            # 6) Auto‑ajustar ancho de columnas
            for column_cells in ws.columns:
                max_len = max(
                    len(str(cell.value)) if cell.value is not None else 0
                    for cell in column_cells
                )
                col_letter = get_column_letter(column_cells[0].column)
                ws.column_dimensions[col_letter].width = max_len + 2

            # 7) Guardar y abrir
            wb.save(ruta)
            if sys.platform.startswith("win"):
                os.startfile(ruta)
            messagebox.showinfo("Éxito", f"Datos exportados correctamente:\n{ruta}", parent=frame)

        except ImportError:
            messagebox.showerror(
                "Error",
                "Para exportar a Excel necesitas instalar openpyxl:\n\npip install openpyxl",
                parent=frame
            )
        except Exception as e:
            messagebox.showerror("Error", f"No se pudo guardar o abrir el archivo:\n{e}", parent=frame)


    btn_exportar = tk.Button(
        frame_botones,
        text="EXPORTAR A EXCEL",
        bg="#3498db",
        fg="white",
        width=22,
        font=("Arial", 9, "bold"),
        command=exportar_a_excel
    )
    btn_exportar.pack(side="left", padx=10)

    # Botón para nueva orden (usando la función ya definida)
    btn_nueva = tk.Button(
        frame_botones,
        text="INGRESAR NUEVA ORDEN",
        bg="lime green",
        fg="white",
        width=22,
        font=("Arial", 9, "bold"),
        command=lambda: agregar_orden_trabajo(actualizar_callback=cargar_datos)
    )
    btn_nueva.pack(side="left", padx=10)

    # ─────────────────────────────────────────────────────────
    # Frame para filtros
    # ─────────────────────────────────────────────────────────

    # Obtener valores únicos de Cliente y Responsable
    with conectar() as conn:
        cur = conn.cursor()
        cur.execute(
            "SELECT DISTINCT cliente "
            "FROM ordenes_trabajo "
            "WHERE cliente IS NOT NULL AND cliente <> ''"
        )
        clientes = [r[0] for r in cur.fetchall()]

        cur.execute(
            "SELECT DISTINCT responsable "
            "FROM ordenes_trabajo "
            "WHERE responsable IS NOT NULL AND responsable <> ''"
        )
        responsables = [r[0] for r in cur.fetchall()]

    # Checkbox para mostrar/ocultar filtros
    var_mostrar_filtros = tk.BooleanVar(value=False)
    
    def toggle_filtros():
        if var_mostrar_filtros.get():
            frame_filtros.grid()
        else:
            frame_filtros.grid_remove()

    chk_filtros = tk.Checkbutton(
        frame_botones,
        text="Mostrar filtros",
        variable=var_mostrar_filtros,
        command=toggle_filtros,
        bg="white"
    )
    chk_filtros.pack(side="right", padx=10)

    # Frame para filtros (sin recuadro)
    frame_filtros = tk.Frame(
        frame,
        bg="white",
        padx=10,
        pady=10
    )
    frame_filtros.grid(row=2, column=0, padx=20, sticky="ew")
    frame.rowconfigure(2, weight=0)
    frame.update_idletasks()
    frame.rowconfigure(2, minsize=frame_filtros.winfo_reqheight())
    frame_filtros.grid_remove()

    vars_filtro = {}
    for idx, (key, label) in enumerate(CAMPOS):
        fila, col = divmod(idx, 5)
        col += 1

        if key == "prioridad":
            fila = 2
            col  = 4

        tk.Label(frame_filtros, text=f"{label}:", bg="white")\
            .grid(row=fila, column=col*2, sticky="e", padx=5, pady=5)
        var = tk.StringVar()
        vars_filtro[key] = var

        def bind_combo(combo, v=var):
            combo.bind("<<ComboboxSelected>>", lambda e: cargar_datos())
            combo.bind("<Delete>",    lambda e: (v.set(""), cargar_datos()))
            combo.bind("<BackSpace>", lambda e: (v.set(""), cargar_datos()))
            combo.bind("<Key>",       lambda e: "break")

        if key == "estado":
            combo = ttk.Combobox(
                frame_filtros,
                textvariable=var,
                values=["Ingresada", "En proceso", "En pausa", "Realizado"],
                width=22,
                state="normal"
            )
            combo.grid(row=fila, column=col*2+1, padx=5, pady=5)
            bind_combo(combo)

        elif key == "prioridad":
            combo = ttk.Combobox(
                frame_filtros,
                textvariable=var,
                values=["Alta", "Media", "Baja"],
                width=22,
                state="normal"
            )
            combo.grid(row=fila, column=col*2+1, padx=5, pady=5)
            bind_combo(combo)

        elif key == "cliente":
            combo = ttk.Combobox(
                frame_filtros,
                textvariable=var,
                values=clientes,
                width=22,
                state="normal"
            )
            combo.grid(row=fila, column=col*2+1, padx=5, pady=5)
            bind_combo(combo)

        elif key == "responsable":
            combo = ttk.Combobox(
                frame_filtros,
                textvariable=var,
                values=responsables,
                width=22,
                state="normal"
            )
            combo.grid(row=fila, column=col*2+1, padx=5, pady=5)
            bind_combo(combo)

        else:
            entry = tk.Entry(
                frame_filtros,
                textvariable=var,
                width=25,
                relief="solid",
                bd=1
            )
            entry.grid(row=fila, column=col*2+1, padx=5, pady=5)
            entry.bind("<KeyRelease>", lambda e: cargar_datos())

    # Filtros de fecha
    lbl_desde = tk.Label(frame_filtros, text="Desde:", bg="white")
    lbl_desde.grid(row=2, column=2, sticky="e", padx=5, pady=5)
    date_desde = DateEntry(frame_filtros, width=18, date_pattern="dd/mm/yyyy")
    date_desde.grid(row=2, column=3, padx=5)
    date_desde.delete(0, "end")

    lbl_hasta = tk.Label(frame_filtros, text="Hasta:", bg="white")
    lbl_hasta.grid(row=2, column=4, sticky="e", padx=5, pady=5)
    date_hasta = DateEntry(frame_filtros, width=18, date_pattern="dd/mm/yyyy")
    date_hasta.grid(row=2, column=5, padx=5)
    date_hasta.delete(0, "end")

    date_desde.bind("<<DateEntrySelected>>", lambda e: cargar_datos())
    date_hasta.bind("<<DateEntrySelected>>", lambda e: cargar_datos())


    # Botones de filtros
    btn_buscar = tk.Button(
        frame_filtros,
        text="Buscar",
        bg="#2980b9",
        fg="white",
        width=15,
        command=cargar_datos
    )
    btn_buscar.grid(row=3, column=4, padx=5)

    btn_limpiar = tk.Button(
        frame_filtros,
        text="Limpiar",
        bg="#e74c3c",
        fg="white",
        width=15,
        command=lambda: (
            [v.set("") for v in vars_filtro.values()],
            date_desde.delete(0, "end"),
            date_hasta.delete(0, "end"),
            cargar_datos()
        )
    )
    btn_limpiar.grid(row=3, column=5, padx=5)


    # Frame para el Treeview
    frame_tree = tk.Frame(frame, bg="white")
    frame_tree.grid(row=3, column=0, sticky="nsew", padx=20, pady=(0, 10))
    frame.rowconfigure(3, weight=1)
    
    # Treeview y scrollbars
    vsb = ttk.Scrollbar(frame_tree, orient="vertical")
    hsb = ttk.Scrollbar(frame_tree, orient="horizontal")
    
    tree = ttk.Treeview(
        frame_tree,
        columns=COLUMNAS,
        show="headings",
        yscrollcommand=vsb.set,
        xscrollcommand=hsb.set,
        selectmode="browse"
    )
    
    vsb.config(command=tree.yview)
    hsb.config(command=tree.xview)
    
    vsb.pack(side="right", fill="y")
    hsb.pack(side="bottom", fill="x")
    tree.pack(fill="both", expand=True)

    # Configurar encabezados
    headers = {campo[0]: campo[1].upper() for campo in CAMPOS}
    widths = {
        "nro_ot": 80,
        "cliente": 150,
        "trabajo": 200,
        "estado": 100,
        "prioridad": 90,
        "responsable": 120,
        "observaciones": 250,
        "fecha_ingreso": 120,
        "fecha_estimada": 120
    }

    for col in COLUMNAS:
        tree.heading(col, text=headers.get(col, col.upper()))
        tree.column(col, width=widths.get(col, 100), anchor="center")

    # Función para ordenar columnas
    def sort_column(col, reverse):
        data = [(tree.set(child, col), child) for child in tree.get_children('')]
        data.sort(reverse=reverse)
        
        for index, (val, child) in enumerate(data):
            tree.move(child, '', index)
        
        tree.heading(col, command=lambda: sort_column(col, not reverse))

    # Configurar ordenamiento
    for col in COLUMNAS:
        tree.heading(col, command=lambda c=col: sort_column(c, False))

    # Función para manejar doble click (modificar orden)
    def on_doble_click(event):
        item_seleccionado = tree.selection()
        if not item_seleccionado:
            return
            
        # Obtener los valores de la fila seleccionada
        valores = tree.item(item_seleccionado[0])['values']
        nro_ot = valores[0]  # Asumimos que nro_ot es la primera columna
        
        # Abrir formulario de modificación
        abrir_formulario_modificar_ot(
            nro_ot=nro_ot,
            parent=frame,
            actualizar_callback=cargar_datos
        )

    # Vincular evento de doble click
    tree.bind("<Double-1>", on_doble_click)

    # Cargar datos iniciales
    cargar_datos()